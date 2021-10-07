# print(sheet.cell_value(0,0))
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import asyncio
from db import database

class excel_input():

    def __init__(self,dbObject):
        self.dbObj = dbObject

    async def getStudents(self,sheet):
        counter = 0 # Counter
        max_rows = sheet.max_row
        check = True

        for i in range(2,max_rows+1):
        # for i in range(2,3):
            cell_obj = sheet.cell(row=i,column=2)
            student_no = cell_obj.value

            if (check):
                courseCode = sheet.cell(row=i,column=1).value
                check = False

            if student_no != None:
                await self.postToEnrolls(student_no,courseCode,self.dbObj)
                counter += 1
            elif student_no == None:
                check = True


    async def postToDB(self):
        await print("post to DB")

    def getCourse(self,sheet):
        max_rows = sheet.max_row
        for i in range(2,max_rows+1):
            cell_obj = sheet.cell(row=i,column=1)
            student_no = cell_obj.value
            print(student_no)

    async def createEnrollsID(self,student_id,code):
        return code[3:]+"_"+student_id
        # Gets the course names from the first column
    
    def createScheduleID(self,code,start,time):
        date_portion = str(start)[:10]
        time_portion = str(time)[:2]
        code_portion = str(code)[3:]

        schedule_id = code_portion+"-"+date_portion+"-"+time_portion
        return schedule_id

    def createDateTimeObj(self,date,time):
        date = str(date)[:11]
        dateTime = date+str(time)
        return dateTime


    async def postToEnrolls(self,student_no,code,dbObj):
        enrollsID = await self.createEnrollsID(student_no,code)
        query = "INSERT INTO enrolls(enrolls_id,student_no,code) VALUES('{}','{}','{}');".format(enrollsID,student_no,code)
        cursor = await dbObj.insertData(query)
        # print(query)
    
    def getSheet(self,name):
        """Gets the active sheet from the excel file

        :param name: [name of excel file]
        :type name: [string]
        :return: [Excel sheet that has data]
        :rtype: [worksheet]
        """
        workBook = load_workbook(name)
        sheet = workBook.active
        return sheet
    
    async def readLabSchedule(self,name,dbObj):
        sheet = self.getSheet(name)
        max_rows = sheet.max_row
        max_columns = sheet.max_column
        # print(max_rows)
        # print(max_columns)
        moreValues = None
        schedule_list = [] # List to be used to create register ID
        for i in range(3,max_rows+1):
            for j in range(2,max_columns+1):
                
                if j == 2:
                    date = sheet.cell(row=i, column=j).value
                    # print(date)
                
                elif j == 3:
                    start_time = sheet.cell(row=i, column=j).value
                    # print(start_time)
                
                elif j == 4:
                    end_time = sheet.cell(row=i, column=j).value
                    # print(end_time)
                
                elif j == 5:
                    course_code = sheet.cell(row=i,column=j).value
                    # print(course_code)

                elif j == 6:
                    activity = sheet.cell(row=i, column=j).value
                    # print(activity) 
            
            start = self.createDateTimeObj(date,start_time)
            end = self.createDateTimeObj(date,end_time)
            schedule_id = self.createScheduleID(course_code,date,start_time)
            if i == 3:
                query = "INSERT INTO lab_schedule(schedule_id,start,end,code,activity) VALUES('{}','{}','{}','{}','{}');".format(schedule_id,start,end,course_code,activity)
            else:
                moreValues = "('{}','{}','{}','{}','{}');".format(schedule_id,start,end,course_code,activity)
        
            query = self.gatherInsertQueries(query,moreValues)
            schedule_list.append((schedule_id,course_code))
            # print(query)
            # print("")
        
        # Add all the schedules to the database
        await dbObj.insertData(query)
        # cursor.execute(query)
        await self.generateRegister(schedule_list,dbObj)

    def gatherInsertQueries(self,query,newValues=None):
        """Combine multiple insert queries into one query,
        In order to allow one call to the db instead of multiple calls

        :param query: The first values to be inserted in the table, (Leading statements of the insert statements)
        :type query: [String]
        :param newValues: [The additional values that have to the leading statement of the query], defaults to None
        :type newValues: [String], optional
        :return: [The new query that has old values and the new values]
        :rtype: [String]
        """
        if(newValues is not None):
            query = query[:-1] + ", "+ newValues
        else:
            query = query
        return query

        
    def generateRegisterId(self,student_no,schedule_id):
        return student_no+"_"+schedule_id

    async def generateRegister(self,schedule_list,dbObj):
        
        for schedule in schedule_list:
            course_code = schedule[1]
            schedule_id = schedule[0] 

            # Get all students
            cursor = dbObj.getDB()
            query = "SELECT * FROM enrolls WHERE code = '{}';".format(course_code)
            cursor.execute(query)

            # Initialise values to be used in loop
            first = True
            moreValues = None

            for entry in cursor:
                student_no = entry[1]
                status = 0
                registerId = self.generateRegisterId(student_no,schedule_id)

                if (first):
                    query = "INSERT INTO register(register_id,status,student_no,schedule_id) VALUES('{}','{}','{}','{}');".format(registerId,status,student_no,schedule_id)
                    first = False
                else:
                    moreValues = "('{}','{}','{}','{}');".format(registerId,status,student_no,schedule_id)

                query = self.gatherInsertQueries(query,moreValues)
           
            await dbObj.insertData(query)







# Initialisations
dbObj = database("eee4022sdatabase-do-user-9871310-0.b.db.ondigitalocean.com",
    "admin",
    "aGAPX1Hn5TdTE-4I",
    "lab_system",
    "25060",
    "mysql_native_password"
    )

wb_obj = load_workbook('course_students.xlsx')
sheet = wb_obj.active
exObj = excel_input(dbObj)

loop = asyncio.get_event_loop()
# loop.run_until_complete(exObj.getStudents(sheet))
loop.run_until_complete(exObj.readLabSchedule('lab_schedule.xlsx',dbObj))

# loop.run_until_complete(exObj.generateRegister('EEE3089F','2045-2020-01-02',dbObj))
# getCourse(sheet)